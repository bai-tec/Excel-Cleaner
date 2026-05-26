import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime
import re
from io import BytesIO
from openpyxl import load_workbook

def load_excel(file):
    try:
        wb = load_workbook(file)
        ws = wb.active
        
        merged_data = []
        for merged_range in list(ws.merged_cells.ranges):
            top_left_value = ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
            merged_data.append({
                'range': merged_range,
                'value': top_left_value
            })
        
        for merged in merged_data:
            ws.unmerge_cells(str(merged['range']))
            mr = merged['range']
            for row in range(mr.min_row, mr.max_row + 1):
                for col in range(mr.min_col, mr.max_col + 1):
                    ws.cell(row=row, column=col).value = merged['value']
        
        temp_file = BytesIO()
        wb.save(temp_file)
        temp_file.seek(0)
        
        df = pd.read_excel(temp_file)
        return df
    except Exception as e:
        st.error(f"读取文件失败: {e}")
        return None

def remove_duplicates(df):
    original_count = len(df)
    df_cleaned = df.drop_duplicates()
    removed_count = original_count - len(df_cleaned)
    return df_cleaned, removed_count

def remove_empty_rows(df):
    original_count = len(df)
    df_cleaned = df.dropna(how='all')
    removed_count = original_count - len(df_cleaned)
    return df_cleaned, removed_count

def remove_empty_columns(df):
    original_count = len(df.columns)
    df_cleaned = df.dropna(axis=1, how='all')
    removed_count = original_count - len(df_cleaned.columns)
    return df_cleaned, removed_count

def fill_missing_values(df, column, method):
    df_filled = df.copy()
    if method == 'forward':
        df_filled[column] = df_filled[column].fillna(method='ffill')
    elif method == 'backward':
        df_filled[column] = df_filled[column].fillna(method='bfill')
    elif method == 'mean':
        if pd.api.types.is_numeric_dtype(df_filled[column]):
            df_filled[column] = df_filled[column].fillna(df_filled[column].mean())
    elif method == 'median':
        if pd.api.types.is_numeric_dtype(df_filled[column]):
            df_filled[column] = df_filled[column].fillna(df_filled[column].median())
    elif method == 'mode':
        df_filled[column] = df_filled[column].fillna(df_filled[column].mode()[0])
    return df_filled

def convert_column_type(df, column, dtype):
    df_converted = df.copy()
    try:
        if dtype == 'numeric':
            df_converted[column] = pd.to_numeric(df_converted[column], errors='coerce')
        elif dtype == 'datetime':
            df_converted[column] = pd.to_datetime(df_converted[column], errors='coerce')
        elif dtype == 'string':
            df_converted[column] = df_converted[column].astype(str)
        return df_converted, True
    except Exception as e:
        return df, False

def validate_phone(phone):
    if pd.isna(phone):
        return False
    phone = str(phone)
    phone = re.sub(r'[^0-9]', '', phone)
    return len(phone) == 11 and phone.startswith('1')

def validate_email(email):
    if pd.isna(email):
        return False
    email = str(email)
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return re.match(pattern, email) is not None

def get_data_summary(df):
    summary = {
        'rows': len(df),
        'columns': len(df.columns),
        'missing': df.isna().sum().sum(),
        'duplicates': df.duplicated().sum(),
        'memory': f'{df.memory_usage(deep=True).sum() / 1024 / 1024:.2f} MB'
    }
    return summary

def main():
    st.set_page_config(page_title="Excel数据清洗工具", page_icon="📊", layout="wide")
    
    st.title("📊 Excel数据清洗工具")
    st.markdown("简单易用的Excel数据清洗和预处理工具")
    
    if 'df' not in st.session_state:
        st.session_state.df = None
    if 'df_cleaned' not in st.session_state:
        st.session_state.df_cleaned = None
    if 'history' not in st.session_state:
        st.session_state.history = []
    
    uploaded_file = st.file_uploader("上传Excel文件", type=['xlsx', 'xls'])
    
    if uploaded_file is not None:
        if st.session_state.df is None:
            st.session_state.df = load_excel(uploaded_file)
            if st.session_state.df is not None:
                st.session_state.df = st.session_state.df.convert_dtypes()
                st.session_state.df_cleaned = st.session_state.df.copy()
        
        if st.session_state.df is not None:
            tab1, tab2, tab3, tab4 = st.tabs(["数据预览", "数据清洗", "数据验证", "导出数据"])
            
            with tab1:
                st.subheader("原始数据")
                st.dataframe(st.session_state.df.head(50), use_container_width=True)
                
                col1, col2, col3 = st.columns(3)
                summary = get_data_summary(st.session_state.df_cleaned)
                
                with col1:
                    st.metric("总行数", summary['rows'])
                with col2:
                    st.metric("总列数", summary['columns'])
                with col3:
                    st.metric("缺失值", summary['missing'])
                
                st.subheader("数据类型信息")
                dtype_info = pd.DataFrame({
                    '列名': st.session_state.df_cleaned.columns,
                    '数据类型': [str(dtype) for dtype in st.session_state.df_cleaned.dtypes.values],
                    '非空值数量': st.session_state.df_cleaned.count().values.tolist(),
                    '缺失值数量': st.session_state.df_cleaned.isna().sum().values.tolist()
                })
                st.dataframe(dtype_info)
                
                st.subheader("数值列统计")
                numeric_cols = st.session_state.df_cleaned.select_dtypes(include=[np.number]).columns
                if len(numeric_cols) > 0:
                    st.dataframe(st.session_state.df_cleaned[numeric_cols].describe())
            
            with tab2:
                st.subheader("数据清洗操作")
                
                col1, col2 = st.columns(2)
                
                with col1:
                    st.write("快速操作")
                    if st.button("删除重复行"):
                        st.session_state.df_cleaned, removed = remove_duplicates(st.session_state.df_cleaned)
                        st.success(f"删除了 {removed} 行重复数据")
                        st.session_state.history.append(f"删除重复行: -{removed}行")
                    
                    if st.button("删除空行"):
                        st.session_state.df_cleaned, removed = remove_empty_rows(st.session_state.df_cleaned)
                        st.success(f"删除了 {removed} 行空数据")
                        st.session_state.history.append(f"删除空行: -{removed}行")
                    
                    if st.button("删除空列"):
                        st.session_state.df_cleaned, removed = remove_empty_columns(st.session_state.df_cleaned)
                        st.success(f"删除了 {removed} 列空数据")
                        st.session_state.history.append(f"删除空列: -{removed}列")
                
                with col2:
                    st.write("列操作")
                    columns = st.session_state.df_cleaned.columns.tolist()
                    
                    if columns:
                        selected_col = st.selectbox("选择列", columns)
                        
                        operation = st.selectbox("操作", ["填充缺失值", "转换数据类型", "删除列"])
                        
                        if operation == "填充缺失值":
                            fill_method = st.selectbox("填充方法", ["前向填充", "后向填充", "平均值", "中位数", "众数"])
                            method_map = {
                                "前向填充": "forward",
                                "后向填充": "backward",
                                "平均值": "mean",
                                "中位数": "median",
                                "众数": "mode"
                            }
                            
                            if st.button("应用填充"):
                                st.session_state.df_cleaned = fill_missing_values(
                                    st.session_state.df_cleaned, 
                                    selected_col, 
                                    method_map[fill_method]
                                )
                                st.success(f"已填充 {selected_col} 列的缺失值")
                                st.session_state.history.append(f"填充 {selected_col}: {fill_method}")
                        
                        elif operation == "转换数据类型":
                            dtype_option = st.selectbox("目标类型", ["数值型", "日期型", "文本型"])
                            dtype_map = {
                                "数值型": "numeric",
                                "日期型": "datetime",
                                "文本型": "string"
                            }
                            
                            if st.button("转换类型"):
                                st.session_state.df_cleaned, success = convert_column_type(
                                    st.session_state.df_cleaned,
                                    selected_col,
                                    dtype_map[dtype_option]
                                )
                                if success:
                                    st.success(f"已将 {selected_col} 转换为 {dtype_option}")
                                    st.session_state.history.append(f"类型转换 {selected_col}: {dtype_option}")
                                else:
                                    st.error("转换失败")
                        
                        elif operation == "删除列":
                            if st.button(f"删除 {selected_col} 列"):
                                st.session_state.df_cleaned = st.session_state.df_cleaned.drop(columns=[selected_col])
                                st.success(f"已删除 {selected_col} 列")
                                st.session_state.history.append(f"删除列: {selected_col}")
                
                st.subheader("清洗后的数据")
                st.dataframe(st.session_state.df_cleaned.head(50))
                
                if st.button("重置数据"):
                    st.session_state.df_cleaned = st.session_state.df.copy()
                    st.session_state.history = []
                    st.info("数据已重置")
            
            with tab3:
                st.subheader("数据验证")
                
                columns = st.session_state.df_cleaned.columns.tolist()
                if columns:
                    validate_col = st.selectbox("选择要验证的列", columns, key="validate_col_tab3")
                    validate_type = st.selectbox("验证类型", ["手机号", "邮箱"], key="validate_type_tab3")
                    
                    if 'validation_result' not in st.session_state:
                        st.session_state.validation_result = None
                    
                    if st.button("开始验证", key="validate_button"):
                        df_validation = st.session_state.df_cleaned.copy()
                        
                        if validate_type == "手机号":
                            df_validation['验证结果'] = df_validation[validate_col].apply(
                                lambda x: "有效" if validate_phone(x) else "无效"
                            )
                        elif validate_type == "邮箱":
                            df_validation['验证结果'] = df_validation[validate_col].apply(
                                lambda x: "有效" if validate_email(x) else "无效"
                            )
                        
                        st.session_state.validation_result = {
                            'df': df_validation,
                            'col': validate_col,
                            'valid_count': (df_validation['验证结果'] == "有效").sum(),
                            'invalid_count': (df_validation['验证结果'] == "无效").sum()
                        }
                    
                    if st.session_state.validation_result is not None:
                        col1, col2 = st.columns(2)
                        with col1:
                            st.metric("有效数据", st.session_state.validation_result['valid_count'])
                        with col2:
                            st.metric("无效数据", st.session_state.validation_result['invalid_count'])
                        
                        st.subheader("验证结果")
                        res_df = st.session_state.validation_result['df']
                        res_col = st.session_state.validation_result['col']
                        st.dataframe(res_df[[res_col, '验证结果']])
            
            with tab4:
                st.subheader("导出数据")
                
                if len(st.session_state.history) > 0:
                    st.write("操作历史:")
                    for i, action in enumerate(st.session_state.history, 1):
                        st.text(f"{i}. {action}")
                
                filename = st.text_input("文件名", "cleaned_data")
                
                output_format = st.selectbox("导出格式", ["Excel (.xlsx)", "CSV"])
                
                if st.button("生成下载"):
                    try:
                        if output_format == "Excel (.xlsx)":
                            output = BytesIO()
                            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                                st.session_state.df_cleaned.to_excel(writer, index=False, sheet_name='清洗后数据')
                                if len(st.session_state.history) > 0:
                                    pd.DataFrame({'操作': st.session_state.history}).to_excel(
                                        writer, index=False, sheet_name='操作记录'
                                    )
                            output.seek(0)
                            st.download_button(
                                label="下载Excel文件",
                                data=output,
                                file_name=f"{filename}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                        else:
                            csv = st.session_state.df_cleaned.to_csv(index=False)
                            st.download_button(
                                label="下载CSV文件",
                                data=csv,
                                file_name=f"{filename}.csv",
                                mime="text/csv"
                            )
                        st.success("文件准备就绪，点击下载")
                    except Exception as e:
                        st.error(f"导出失败: {e}")

if __name__ == "__main__":
    main()
